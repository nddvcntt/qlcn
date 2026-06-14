#!/usr/bin/env python3
"""Convert Excel -> JSON for sheets we need."""
import json
import os
import openpyxl
import sys

EXCEL = r"D:\Other\QLCN\Untitled spreadsheet.xlsx"
OUT = r"D:\Other\QLCN\qlcn-app\scripts\excel-data.json"

wb = openpyxl.load_workbook(EXCEL, data_only=True)
data: dict = {}
for name in wb.sheetnames:
    ws = wb[name]
    rows = []
    for row in ws.iter_rows(values_only=True):
        # Convert dates to ISO
        row_data = []
        for v in row:
            if hasattr(v, "isoformat"):
                row_data.append(v.isoformat())
            else:
                row_data.append(v)
        # Skip totally empty rows
        if any(v is not None and v != "" for v in row_data):
            rows.append(row_data)
    data[name] = rows

os.makedirs(os.path.dirname(OUT), exist_ok=True)
with open(OUT, "w", encoding="utf-8") as f:
    json.dump(data, f, ensure_ascii=False, indent=2, default=str)

print(f"Saved {OUT}")
for name, rows in data.items():
    sys.stdout.buffer.write(f"  {name}: {len(rows)} rows\n".encode("utf-8"))
